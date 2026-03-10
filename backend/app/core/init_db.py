"""
Database initialization and Excel import.

Imports the client's existing Excel spreadsheet into SQLite,
mapping columns and color-coded statuses to the new schema.
"""
import asyncio
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from app.core.database import engine, Base, async_session
from app.models.models import Cliente


# Map Excel fill colors to client status
COLOR_STATUS_MAP = {
    "FFFF0000": "no_llamar",     # Red
    "FF00FF00": "venta",          # Green
    "FFFFFF00": "equivocado",     # Yellow
    "FF800080": "cita",           # Purple
    "FFCC00CC": "cita",           # Another purple variant
    "FF9900CC": "cita",           # Purple variant
    "FF0000FF": "seguimiento",    # Blue
    "FF0066FF": "seguimiento",    # Blue variant
}


def detect_status_from_color(cell) -> str:
    """Detect client status from Excel cell background color."""
    fill = cell.fill
    if fill and fill.start_color and fill.start_color.rgb:
        color = str(fill.start_color.rgb)
        if color in COLOR_STATUS_MAP:
            return COLOR_STATUS_MAP[color]
        # Check if it looks like purple (the most common status in the sheet)
        if "CC" in color or "80" in color or "99" in color:
            return "cita"
    return "nuevo"


def import_excel(filepath: str) -> list[dict]:
    """
    Parse the existing Excel spreadsheet and return client records.

    Expected columns (based on screenshot):
      A: Row number
      B: Nombre y apellido
      C: Teléfono
      D: Fuente
      E: Zona
      F: Dirección
      G: Notas del vendedor
      H: Notas del Telemarketing
      I: Resultados
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    clients = []
    header_row = 8  # Based on the screenshot, headers are in row 8

    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        # Skip empty rows
        nombre_cell = row[1]  # Column B
        if not nombre_cell.value:
            continue

        nombre = str(nombre_cell.value).strip()
        telefono = str(row[2].value).strip() if row[2].value else ""

        if not telefono:
            continue

        # Detect status from row color
        estado = detect_status_from_color(nombre_cell)

        client = {
            "nombre_apellido": nombre,
            "telefono": telefono,
            "fuente": str(row[3].value).strip() if row[3].value else None,
            "zona": str(row[4].value).strip() if row[4].value else None,
            "direccion": str(row[5].value).strip() if row[5].value else None,
            "estado": estado,
        }

        # Add notes as a first "virtual" visit note if present
        notas_vendedor = str(row[6].value).strip() if row[6].value else None
        notas_telemarketing = str(row[7].value).strip() if row[7].value else None
        resultados = str(row[8].value).strip() if row[8].value else None

        client["_notas_vendedor"] = notas_vendedor
        client["_notas_telemarketing"] = notas_telemarketing
        client["_resultados"] = resultados

        clients.append(client)

    wb.close()
    return clients


async def init_database(excel_path: str = None):
    """
    Initialize the database and optionally import from Excel.

    Usage:
        python -m app.core.init_db                    # Just create tables
        python -m app.core.init_db path/to/excel.xlsx  # Create + import
    """
    # Create all tables
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    print(f"Database tables created.")

    if excel_path:
        print(f"Importing from Excel: {excel_path}")
        clients = import_excel(excel_path)

        async with async_session() as session:
            imported = 0
            skipped = 0

            for client_data in clients:
                # Pop non-model fields
                client_data.pop("_notas_vendedor", None)
                client_data.pop("_notas_telemarketing", None)
                client_data.pop("_resultados", None)

                # Check if phone already exists
                existing = await session.execute(
                    select(Cliente).where(Cliente.telefono == client_data["telefono"])
                )
                if existing.scalar_one_or_none():
                    skipped += 1
                    continue

                cliente = Cliente(**client_data)
                session.add(cliente)
                imported += 1

            await session.commit()
            print(f"Import complete: {imported} created, {skipped} skipped (duplicate phone)")

    print("Database initialization complete.")


if __name__ == "__main__":
    excel_file = sys.argv[1] if len(sys.argv) > 1 else None
    asyncio.run(init_database(excel_file))
